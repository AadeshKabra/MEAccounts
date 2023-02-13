
from flask import Flask, render_template, request, flash, send_file, redirect, url_for, session, send_from_directory
import pandas as pd
from flask_pymongo import PyMongo, pymongo

from datetime import datetime, date
import matplotlib.pyplot as plt

from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
from io import BytesIO
import os
import openpyxl

# import schedule
# from openpyxl import load_workbook
from pymongo import MongoClient
import datetime
# from webui import WebUI
# from flaskwebgui import FlaskUI

# from webui import WebUI


app = Flask(__name__)
app.config["SECRET_KEY"] = "013a658d9c8323f8e0af1f8ba8b4bcf30456a4c0"
app.config["MONGO_URI"] = "mongodb+srv://sagar:sagar@cluster0.gwnjuq6.mongodb.net/?retryWrites=true&w=majority"

# ui = WebUI(app, debug=True)
# ui = FlaskUI(app)

employee_names = []
# production_items = []

current_month = datetime.datetime.now().strftime("%B")
current_year = datetime.datetime.now().year

mongo_client = pymongo.MongoClient("mongodb+srv://sagar:sagar@cluster0.gwnjuq6.mongodb.net/?retryWrites=true&w=majority", maxPoolSize=50, connect=False)
db = pymongo.database.Database(mongo_client, 'ME')
users = pymongo.collection.Collection(db, 'users')

# app.config["MONGO_URI"] = "mongodb://127.0.0.1:27017/ME"
# mongo_client = PyMongo(app)
# app.secret_key = 'your_secret_key'
# db = mongo_client.ME
# db = mongo_client.get_database('ME')

# user_collection = PyMongo.collection.Collection(db, 'users')
# users = mongo_client.db.users



# Raw Material Section


# Displaying raw material page
@app.route("/raw_material", methods=['GET'])
def raw_material():
    return render_template("rawMaterial.html")


# Displaying page for insertion of raw material
@app.route("/insert_raw", methods=["POST"])
def insert():
    return render_template("insert.html")


# Insertion of raw material
@app.route("/insert_raw_material", methods=['POST', 'GET'])
def insert_raw():
    # rate = request.form.get("rate")
    # print(rate)
    app.config["MONGO_URI"] = "mongodb://127.0.0.1:27017/ME"
    # mongo = PyMongo(app)
    raw = mongo_client.db.raw_material
    element = {
        "name": request.form.get("raw_name"),
        "units": request.form.get("units"),
        "opening": request.form.get("opening"),
        "received": request.form.get("received"),
        "rate": request.form.get("rate")
    }
    document = raw.find_one({"name": request.form.get("raw_name")})
    if document:
        print("Raw Material exists")
        # flash("Raw Material exists, please try updating it!")
        return render_template("rawMaterial.html", insertMessage=True)
        # return redirect(request.referrer)
        # return redirect(url_for('raw_material'))
    else:
        raw.insert_one(element)
        return render_template("index.html")
        # return redirect(request.referrer)
        # return redirect(url_for('raw_material'))

    return render_template("index.html")


# Displaying page for updating of raw material
@app.route("/update_raw", methods=['POST'])
def update():
    list_names = []
    raw = mongo_client.db.raw_material
    documents = raw.find({})
    for i in documents:
        list_names.append(i["name"])
    # print(list_names)
    return render_template("update.html", raw_materials=list_names)


# Updating the raw material
@app.route("/update_raw_material", methods=['POST', 'GET'])
def update_raw():
    app.config["MONGO_URI"] = "mongodb://127.0.0.1:27017/ME"
    # mongo = PyMongo(app)
    raw = mongo_client.db.raw_material
    u_name = request.form.get("item")
    u_received = request.form.get("received")
    u_rate = request.form.get("rate")
    original = list(raw.find({
        "name": {"$eq": u_name}
    }))
    received = int(u_received) + int(original[0]['received'])
    rate = float(float(u_rate) + float(original[0]['rate'])) / 2
    print(original[0])
    raw.find_one_and_update({"name": u_name}, {"$set": {"received": received, "rate": rate}})

    return render_template("index.html")


# Displaying the report of raw material
@app.route("/open_raw", methods=['GET'])
def view():
    # list_issued = [0, 0, 0, 0, 761.83, 0, 0, 0]
    app.config["MONGO_URI"] = "mongodb://127.0.0.1:27017/ME"
    # mongo = PyMongo(app)
    raw = mongo_client.db.raw_material
    items = mongo_client.db.prodItems
    production = mongo_client.db.production

    groups = {}
    documents = items.find({})
    for doc in documents:
        raw_material = doc['RawMaterial']
        name = doc['Name']

        if raw_material in groups:
            groups[raw_material].append(name)
        else:
            groups[raw_material] = [name]

    list_quantity = {}
    for i in groups:
        # print(i)
        total_quantity = 0
        for j in groups[i]:
            document = production.find_one({"Name": j})
            # print(document)
            total_quantity += float(document['Kgs'])
        list_quantity[i] = total_quantity
    print("List Quantity", list_quantity)

    names = list(raw.find({}, {"name": 1, "_id": 0}))
    list_names = []
    for i in names:
        list_names.append(i['name'])

    list_issued = []
    for i in list_names:
        if i not in list_quantity:
            list_issued.append(0)
        else:
            list_issued.append(list_quantity[i])

    units = list(raw.find({}, {"units": 1, "_id": 0}))
    list_units = []
    for i in units:
        list_units.append(i['units'])

    opening = list(raw.find({}, {"opening": 1, "_id": 0}))
    list_opening = []
    for i in opening:
        list_opening.append(float(i['opening']))

    received = list(raw.find({}, {"received": 1, "_id": 0}))
    list_received = []
    for i in received:
        list_received.append(float(i['received']))

    rate = list(raw.find({}, {"rate": 1, "_id": 0}))
    list_rate = []
    for i in rate:
        list_rate.append(float(i['rate']))

    bal = []
    for i in range(len(list_opening)):
        bal.append(float(list_opening[i]) + float(list_received[i]) - float(list_issued[i]))

    opening_value = []
    received_value = []
    issued_value = []
    for i in range(len(list_rate)):
        opening_value.append(float(list_opening[i]) * float(list_rate[i]))
        received_value.append(float(list_received[i]) * float(list_rate[i]))
        issued_value.append(float(list_issued[i]) * float(list_rate[i]))

    data = pd.DataFrame(list(zip(list_names, list_units, list_opening, list_received, list_issued, bal, list_rate, opening_value, received_value, issued_value)), columns=['Description', 'UNITS', 'OPENING', 'RECEIVED', 'ISSUED', 'BALANCE', 'RATE', 'OPENING by Value', 'RECEIVED by Value', 'ISSUED by Value'])

    data['BALANCE by Value'] = data['BALANCE']*data['RATE']
    # data.append(data.sum(numeric_only=True), ignore_index=True)
    # data.loc[:, 'Row_Total'] = data.sum(numeric_only=True, axis=1)
    data_total = data.append(data.sum(axis=0), ignore_index=True)

    # Set the name of the last row to "Total"
    data_total.loc[data_total.index[-1], 'Description'] = "Total"
    data_total.loc[data_total.index[-1], 'UNITS'] = ""
    data_total.loc[data_total.index[-1], 'RATE'] = ""
    data_total.style.set_properties(subset=['Description'], **{'font-weight': 'bold'})
    # data_total.style.set_properties(**{'background-color': 'white', 'color': 'blue', 'border-color': 'black'})

    file_path = "static/" + str(current_month) + str(current_year) + ".xlsx"
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        workbook.save(file_path)
        print("File Created successfully")

    with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        data_total.to_excel(writer, 'Raw Material')


    table = data_total.to_html()
    return render_template("view.html", table=table)


# Production section

# Displaying the production page
@app.route("/production", methods=['POST', 'GET'])
def prod():
    return render_template("production.html")


# Displaying page of insertion of production item
@app.route("/insertItem", methods=['POST', 'GET'])
def insertItem():
    raw = mongo_client.db.raw_material
    raw_materials = []
    for i in raw.find({}, {"name": 1, "_id": 0}):
        if i['name']:
            raw_materials.append(i["name"])

    return render_template("insertItem.html", raw_materials=raw_materials)


# Inserting new production item
@app.route("/newProd", methods=['POST'])
def newProd():
    app.config["MONGO_URI"] = "mongodb://127.0.0.1:27017/ME"
    # mongo = PyMongo(app)
    prodItems = mongo_client.db.prodItems
    name = request.form.get("name")
    weight = request.form.get("weight")
    raw = request.form.get("raw")
    rate = request.form.get("rate")
    # opening = request.form.get("opening")
    item = {
        "Name": name,
        "Weight": weight,
        "RawMaterial": raw,
        "Rate": rate
    }
    prodItems.insert_one(item)
    # production_items.append(name)
    items = []
    for i in prodItems.find({}, {"Name": 1}):
        items.append(i['Name'])
    print(items)

    return render_template("updateProd.html", items=items)


# Displaying the page for updation of production item
@app.route("/updateProd", methods=['POST', 'GET'])
def updateProd():
    app.config["MONGO_URI"] = "mongodb://127.0.0.1:27017/ME"
    # mongo = PyMongo(app)
    prodItems = mongo_client.db.prodItems
    production_items = list(prodItems.find({}, {"_id": 0, "Name": 1}))
    list_prod_items = []
    for i in range(len(production_items)):
        list_prod_items.append(production_items[i]['Name'])
    # print(production_items)
    print(list_prod_items)
    return render_template("updateProd.html", items=list_prod_items)


# Updating production item
@app.route("/uProduction", methods=['POST', 'GET'])
def uProduction():
    date = request.form.get("date")
    item = request.form.get("item")
    quantity = request.form.get("prodQuan")
    print(date, item, quantity)
    day = int(date[8] + date[9])
    month = int(date[5] + date[6])
    odd = True
    if month not in [1, 3, 5, 7, 8, 10, 12]:
        odd = False
    print(day, month)
    app.config["MONGO_URI"] = "mongodb://127.0.0.1:27017/ME"
    # mongo = PyMongo(app)
    production = mongo_client.db.production
    items = mongo_client.db.prodItems

    cursor = production.find_one({"Name": item})
    if cursor:
        print("Document exists")
        # print(cursor)
        doc = cursor['Production']
        initial = doc[str(day)]
        count = int(initial) + int(quantity)
        cursor['Production'][str(day)] = count
        print(cursor)
        doc_del = production.find_one({"Name": item})

        doc_item = items.find_one({"Name": item})
        weight = float(doc_item['Weight'])
        total_quantity = 0
        if odd:
            for i in range(1, 32):
                total_quantity += int(cursor['Production'][str(i)])
        else:
            for i in range(1, 31):
                total_quantity += int(cursor['Production'][str(i)])
        kgs = weight * total_quantity
        cursor["Kgs"] = kgs

        production.delete_one(doc_del)
        production.insert_one(cursor)
    else:
        print("Document does not exist")
        doc = {}
        if odd:
            for i in range(1, 32):
                if day == i:
                    doc[str(i)] = quantity
                else:
                    doc[str(i)] = 0
        else:
            for i in range(1, 31):
                if day == i:
                    doc[str(i)] = quantity
                else:
                    doc[str(i)] = 0

        document = {
            "Name": item,
            "Production": doc
        }

        doc_item = items.find_one({"Name": item})
        weight = float(doc_item['Weight'])
        total_quantity = 0
        if odd:
            for i in range(1, 32):
                total_quantity += int(document['Production'][str(i)])
        else:
            for i in range(1, 31):
                total_quantity += int(document['Production'][str(i)])

        print("Weight", weight)
        print("Quantity", total_quantity)
        kgs = weight * total_quantity
        print("Kgs", kgs)
        document["Kgs"] = kgs
        production.insert_one(document)

    return render_template("index.html")


# Displaying report of production
@app.route("/viewProd", methods=['POST', 'GET'])
def viewProduction():
    print("Production report")
    app.config["MONGO_URI"] = "mongodb://127.0.0.1:27017/ME"
    # mongo = PyMongo(app)
    raw_material = mongo_client.db.raw_material
    items = mongo_client.db.prodItems
    production = mongo_client.db.production

    documents = raw_material.find({}, {"name": 1})
    raw_materials = []
    for doc in documents:
        raw_materials.append(doc['name'])
    print(raw_materials)

    groups = {}
    documents = items.find({})
    for doc in documents:
        raw_material = doc['RawMaterial']
        name = doc['Name']

        if raw_material in groups:
            groups[raw_material].append(name)
        else:
            groups[raw_material] = [name]

    print(groups)
    list_quantity = {}
    for i in groups:
        # print(i)
        total_quantity = 0
        for j in groups[i]:
            document = production.find_one({"Name": j})
            # print(document)
            total_quantity += float(document['Kgs'])
        list_quantity[i] = total_quantity
    print("List Quantity", list_quantity)

    raw_material_data = []
    prod_items = []
    prod_docs = production.find({}, {"Name": 1, "_id": 0})
    for prod_doc in prod_docs:
        prod_items.append(prod_doc['Name'])
    # print("Prod Items", prod_items)
    for i in prod_items:
        doc = items.find_one({"Name": i})
        raw_material_data.append(doc['RawMaterial'])

    # for i in prod_items:
    #     doc = production.find_one({}, {"Name": i})
    # for i in list_quantity:
    #     raw_material_data.append(i)
    print("Raw Materials", raw_material_data)

    quantity_data = []
    for i in list_quantity:
        quantity_data.append(list_quantity[i])
    # diff = len(raw_material_data) - len(quantity_data)
    # for i in range(diff):
    #     quantity_data.append(None)
    print("Quantities", quantity_data)

    list_items = []
    items = production.find({}, {"Name": 1, "_id": 0})
    for item in items:
        # print(item)
        list_items.append(item['Name'])
    print("Production Items", list_items)

    dic_production = {}
    day_by_day_production = production.find({}, {"Name": 1, "Production": 1, "_id": 0})
    for i in day_by_day_production:
        dic_production[i['Name']] = i['Production']
    print(dic_production)

    kgs = []
    quantities = production.find({}, {"Name": 1, "Kgs": 1, "_id": 0})
    for quantity in quantities:
        kgs.append(quantity['Kgs'])
    print(kgs)

    items = mongo_client.db.prodItems
    weights = []
    for item in list_items:
        item_weight = items.find_one({"Name": item})
        weights.append(item_weight['Weight'])
    print("Weights", weights)
    print("-----------------------------------")
    # print(list_items)

    data1 = pd.DataFrame(list(zip(list_items, kgs, weights, raw_material_data)),
                         columns=['Items', 'Quantity (Kgs)', 'Weight', 'Raw Material used'])

    data2 = pd.DataFrame(dic_production)
    print(dic_production)
    data2 = data2.transpose()
    data2 = data2.reset_index()
    data2.drop(['index'], axis=1, inplace=True)
    data2 = data2.reset_index()
    # data2.rename(columns={'index': 'Items'}, inplace=True)

    data2 = data2.drop(['index'], axis=1)
    data2['Total'] = data2.sum(axis=1)
    # print(data2)


    result = pd.concat([data1, data2], axis=1, join='outer')
    # result = result.drop(result.columns[4], axis=1)
    # result['Total'] = result.sum(axis=1)
    list_total = []
    for i in dic_production:
        total = 0
        # print(dic_production[i])
        for j in dic_production[i]:
            total += float(dic_production[i][j])
        list_total.append(total)
    print(list_total)
    result['Total'] = list_total


    table1 = result.to_html()

    raw = []
    for i in groups:
        raw.append(i)
    print(raw)
    data3 = pd.DataFrame(list(zip(raw, quantity_data)), columns=['Raw Material', 'Quantity of material used'])
    table2 = data3.to_html()

    file_path = "static/" + str(current_month) + str(current_year) + ".xlsx"
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        workbook.save(file_path)
        print("File Created successfully")

    with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        result.to_excel(writer, 'Production')
        # data3.to_excel(writer, 'Production', startrow=len(result)+1, index=False)


    return render_template("prodView.html", report1=table1, report2=table2)


# TRAUB Attendance section

# Displaying employees page
@app.route("/attendanceT", methods=['POST', 'GET'])
def attendanceT():
    return render_template("Employees/employees.html")


# Displaying the page for adding new employee
@app.route("/addE", methods=['POST'])
def addE():
    return render_template("Employees/addEmploy.html")


# Adding new employee
@app.route("/addEmployee", methods=['POST'])
def add_employee():
    app.config["MONGO_URI"] = "mongodb://127.0.0.1:27017/ME"
    # mongo = PyMongo(app)
    employees = mongo_client.db.employees

    name = request.form.get("eName")
    age = request.form.get("eAge")
    number = request.form.get("eNumber")
    rate = request.form.get("eRate")
    details = {
        "Name": name,
        "Age": age,
        "Contact": number,
        "Rate": rate
    }
    employees.insert_one(details)
    global employee_names
    employee_names.append(name)
    return render_template("Employees/employees.html")


# Displaying page for marking the attendance of employee
@app.route("/markE", methods=['POST'])
def markEmployee():
    app.config["MONGO_URI"] = "mongodb://127.0.0.1:27017/ME"
    # mongo = PyMongo(app)
    employees = mongo_client.db.employees
    names = list(employees.find({}, {"Name": 1, "_id": 0}))
    list_names = []
    for i in range(len(names)):
        list_names.append(names[i]['Name'])
    print(list_names)
    return render_template("Employees/markEmployee.html", employees=list_names)


# Marking the attendance of employee
@app.route("/attendance", methods=['POST'])
def attendance():
    date = request.form.get("date")
    hours = request.form.getlist("hours")
    # print(date, hours)
    employees = mongo_client.db.employees
    names = list(employees.find({}, {"Name": 1, "_id": 0}))
    list_names = []
    for i in range(len(names)):
        list_names.append(names[i]['Name'])

    day = int(date[8] + date[9])
    month = int(date[5] + date[6])
    print("Day", day, "Month", month)
    odd = [1, 3, 5, 7, 8, 10, 12]
    attend = mongo_client.db.attendance
    for i in range(len(list_names)):
        a_doc = attend.find_one({"Name": list_names[i]})
        # print("A_doc", a_doc)
        if a_doc:
            print("Document exists")
            # document = {}
            doc_del = attend.find_one({"Name": list_names[i]})
            print(type(doc_del))
            # document["Name"] = doc_del["Name"]
            dic = doc_del["Attendance"]
            dic[str(day)] = int(hours[i])
            attend.update_one({"Name": list_names[i]}, {"$set": {"Attendance": dic}})
        else:
            print("Document does not exist")
            document = {
                "Name": list_names[i],
            }
            doc = {}
            if month in odd:
                for j in range(1, 32):
                    if j == day:
                        doc[str(day)] = int(hours[i])
                    else:
                        doc[str(j)] = 0
            else:
                for j in range(1, 31):
                    if j == day:
                        doc[str(day)] = int(hours[i])
                    else:
                        doc[str(j)] = 0
            document["Attendance"] = doc
            print(document)
            attend.insert_one(document)
    return render_template("Employees/employees.html")


# Displaying the page for removal of employee
@app.route("/removeE", methods=['POST'])
def removeE():
    return render_template("Employees/removeEmployee.html")


# Removing employee
@app.route("/remove", methods=['POST'])
def remove():
    name = request.form.get("name")
    employees = mongo_client.db.employees
    employees.delete_one({"Name": name})
    return render_template("Employees/employees.html")


# Displaying attendance report
@app.route("/viewAttendance", methods=['POST'])
def viewAttendance():
    attendance = mongo_client.db.attendance
    employees = mongo_client.db.employees
    names = list(attendance.find({}, {"Name": 1, "_id": 0}))
    # print(names)
    employee_names = []
    for i in names:
        employee_names.append(i['Name'])
    attend = list(attendance.find({}, {"Attendance": 1, "_id": 0}))
    print(attend)
    data1 = pd.DataFrame(employee_names, columns=['Name of Employee'])
    # print(data1)
    # print(attend[0]['Attendance'])
    # data2 = pd.DataFrame([attend[0]["Attendance"]])
    data2 = pd.DataFrame()
    for i in range(len(employee_names)):
        temp_data = pd.DataFrame([attend[i]["Attendance"]])
        data2 = data2.append(temp_data, ignore_index=True)
    print(data2)

    data = pd.concat([data1, data2], axis=1)
    data['Total'] = data.sum(axis=1)

    rate = list(employees.find({}, {"Rate": 1, "_id": 0}))
    print(rate)
    list_rate = []
    for i in rate:
        list_rate.append(float(i["Rate"]))

    temp_data = pd.DataFrame(list_rate, columns=["Rate"])
    data = pd.concat([data, temp_data], axis=1)
    # data["Rate"]
    data['Days'] = data['Total']/8
    data["Net Amount"] = data["Days"] * data["Rate"]

    file_path = "static/" + str(current_month) + str(current_year) + ".xlsx"
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        workbook.save(file_path)
        print("File Created successfully")

    with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        data.to_excel(writer, 'Attendance TRAUB')

    table = data.to_html()
    return render_template("Employees/viewAttendance.html", table=table)


# Dispatch section


@app.route("/dispatch", methods=['GET'])
def dispatchProduction():
    return render_template("Dispatch/dispatch.html")


@app.route("/dispatchInsert", methods=["GET"])
def dispatchInsert():
    production = mongo_client.db.production
    names = list(production.find({}, {"Name": 1, "_id": 0}))
    list_names = []
    for i in names:
        list_names.append(i["Name"])
    print(list_names)
    return render_template("Dispatch/insertDispatch.html", items=list_names)


flag = True


@app.route("/updateDispatch", methods=['POST'])
def updateDispatch():
    name = request.form.get("item")
    date = request.form.get("date")
    quantity = request.form.get("quantity")
    items = mongo_client.db.prodItems
    dispatch = mongo_client.db.dispatch
    # print(name, date)
    day = int(date[8] + date[9])
    month = int(date[5] + date[6])
    print(day, month)
    odd = [1, 3, 5, 7, 8, 10, 12]

    names = list(items.find({}, {"Name": 1, "_id": 0}))
    list_names = []
    for i in names:
        list_names.append(i['Name'])

    doc = dispatch.find_one({"Name": name})
    if doc:
        # d_doc = dispatch.find({"Name": name})
        print(doc)
        dispatch_dic = doc["Dispatch"]
        initial = dispatch_dic[str(day)]
        count = int(initial) + int(quantity)
        dispatch_dic[str(day)] = count
        dispatch.update_one({"Name": name}, {"$set": {"Dispatch": dispatch_dic}})
    else:
        for i in list_names:
            document = {
                "Name": i,
                "Quantity": 0
            }
            doc = {}
            if month in odd:
                for j in range(1, 32):
                    doc[str(j)] = 0
            else:
                for j in range(1, 31):
                    doc[str(j)] = 0
            document["Dispatch"] = doc

            dispatch.insert_one(document)

        document = dispatch.find_one({"Name": name})
        # document["Dispatch"][str(day)] = quantity
        dispatch_dic = document['Dispatch']
        dispatch_dic[str(day)] = quantity
        dispatch.update_one({"Name": name}, {"$set": {"Dispatch": dispatch_dic}})

    return render_template("Dispatch/dispatch.html")


@app.route("/updateSchedule", methods=['GET'])
def schedule():
    production = mongo_client.db.production
    names = list(production.find({}, {"Name": 1, "_id": 0}))
    list_names = []
    for i in names:
        list_names.append(i['Name'])

    return render_template("Dispatch/schedule.html", items=list_names)


@app.route("/insertSchedule", methods=['POST'])
def insertSchedule():
    name = request.form.get("item")
    quantity = request.form.get("schedule")
    schedule = mongo_client.db.schedule
    document = {
        "Name": name,
        "Schedule": quantity
    }
    schedule.insert_one(document)
    return render_template("Dispatch/dispatch.html")


@app.route("/viewScheduleReport", methods=['post'])
def view_report():
    items = mongo_client.db.prodItems

    # Column of names
    names = list(items.find({}, {"Name": 1, "_id": 0}))
    list_names = []
    for i in names:
        list_names.append(i['Name'])
    # print(list_names)

    # Column of scheduled quantities
    schedule = mongo_client.db.schedule
    list_schedule = []
    for i in list_names:
        doc = schedule.find_one({"Name": i})
        if doc:
            list_schedule.append(int(doc["Schedule"]))
        else:
            list_schedule.append(0)
    # print(list_schedule)

    # Columns of days
    dispatch = mongo_client.db.dispatch
    data1 = pd.DataFrame(list_names, columns=['Item Name'])
    data2 = pd.DataFrame()
    for i in list_names:
        doc = dispatch.find_one({"Name": i})
        # data2 = pd.DataFrame(doc["Dispatch"])
        # print(data2)
        # print(doc['Dispatch'])
        data3 = pd.DataFrame(list(doc['Dispatch'].items()), columns=[1, 2])
        data3 = data3.transpose()
        first_row = data3.iloc[0]
        data3.columns = first_row
        data3 = data3[1:]
        # print(data3)
        data2 = data2.append(data3)

    data1.index = data2.index
    data = pd.concat([data1, data2], axis=1)

    # Column of Total
    total = []
    for i in list_names:
        document = dispatch.find_one({"Name": i})
        count = 0
        for j in document['Dispatch']:
            count += int(document['Dispatch'][j])
        total.append(count)
    # print(total)
    data['Despatched'] = total

    data['Schedule'] = list_schedule

    data['Balance Schedule'] = data['Schedule'] - data['Despatched']

    # PO rate column
    list_po = []
    for i in list_names:
        doc = items.find_one({"Name": i})
        list_po.append(float(doc['Rate']))

    data['PO Rate'] = list_po

    # Total Sale column
    data['Total Sale'] = data['PO Rate'] * data['Despatched']

    file_path = "static/" + str(current_month) + str(current_year) + ".xlsx"
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        workbook.save(file_path)
        print("File Created successfully")

    with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        data.to_excel(writer, 'Dispatch')

    table = data.to_html()

    return render_template("Dispatch/viewReport.html", table=table)


# Creditors Section

# Displaying Creditors Section
@app.route("/creditors", methods=['GET'])
def creditors():
    return render_template("Creditors/creditors.html")


# Displaying page for adding creditors
@app.route("/addCreditor", methods=['POST'])
def addCreditor():
    return render_template("Creditors/addCreditor.html")


# Adding Creditor
@app.route("/aCreditor", methods=['POST'])
def aCreditor():
    name = request.form.get("name")
    date = request.form.get("date")
    amount = request.form.get("amount")
    doc = {
        date: float(amount)
    }
    document = {
        "Name": name,
        "30 Days": float(amount),
        "60 Days": 0,
        "90 Days": 0,
        "90 Onwards": 0,
        "Production": doc
    }
    creditors = mongo_client.db.creditors
    creditors.insert_one(document)
    return render_template("Creditors/creditors.html")


# Displaying page for adding new Due amount
@app.route("/addDue", methods=['POST'])
def addDue():
    list_names = []
    creditors = mongo_client.db.creditors
    names = creditors.find({}, {"Name": 1, "_id": 0})
    for i in names:
        list_names.append(i["Name"])

    return render_template("Creditors/addDue.html", creditors=list_names)


# Adding new Due amount
@app.route("/aDue", methods=['POST'])
def aDue():
    name = request.form.get("name")
    amount = request.form.get("amount")
    date = request.form.get("date")
    creditors = mongo_client.db.creditors
    document = creditors.find_one({"Name": name})
    print(document)
    doc = document['Production']

    today = datetime.today().strftime('%Y-%m-%d')
    # print(today)
    date1 = datetime.strptime(date, '%Y-%m-%d')
    date2 = datetime.strptime(today, '%Y-%m-%d')
    difference = (date1 - date2).days
    # difference = abs(date1 - date2)
    difference = abs(difference)
    print(difference)
    if difference <= 30:
        initial = document["30 Days"]
        count = float(initial) + float(amount)
        creditors.update_one({"Name": name}, {"$set": {"30 Days": count}})
    elif difference <= 60:
        initial = document["60 Days"]
        count = float(initial) + float(amount)
        creditors.update_one({"Name": name}, {"$set": {"60 Days": count}})
    elif difference <= 90:
        initial = document["90 Days"]
        count = float(initial) + float(amount)
        creditors.update_one({"Name": name}, {"$set": {"90 Days": count}})
    else:
        initial = document["90 Onwards"]
        count = float(initial) + float(amount)
        creditors.update_one({"Name": name}, {"$set": {"90 Onwards": count}})


    if date in doc:
        initial = float(doc['Production'][date])
        count = initial + float(amount)
        creditors.update_one({"Name": name}, {"$set": {"Production." + date: count}})
    else:
        creditors.update_one({"Name": name}, {"$set": {"Production." + date: float(amount)}})

    return render_template("Creditors/creditors.html")


# Displaying page for releasing new amount
@app.route("/releaseDue", methods=['POST'])
def releaseDue():
    list_names = []
    creditors = mongo_client.db.creditors
    names = creditors.find({}, {"Name": 1, "_id": 0})
    for i in names:
        list_names.append(i["Name"])
    return render_template("Creditors/releaseDue.html", creditors=list_names)


# Releasing amount
@app.route("/rDue", methods=['POST'])
def rDue():
    name = request.form.get("name")
    amount = request.form.get("amount")
    creditors = mongo_client.db.creditors
    document = creditors.find_one({"Name": name})
    print(document)
    if document["90 Onwards"] > 0:
        initial = document["90 Onwards"]
        count = float(initial) - float(amount)
        if count<0:
            amount = abs(count)
            count = 0
        else:
            amount = 0
        creditors.update_one({"Name": name}, {"$set": {"90 Onwards": count}})
    if document["90 Days"] > 0 and float(amount) > 0:
        initial = document["90 Days"]
        count = float(initial) - float(amount)
        if count < 0:
            amount = abs(count)
            count = 0
        else:
            amount = 0
        creditors.update_one({"Name": name}, {"$set": {"90 Days": count}})
    if document["60 Days"] > 0 and float(amount) > 0:
        initial = document["60 Days"]
        count = float(initial) - float(amount)
        if count < 0:
            amount = abs(count)
            count = 0
        else:
            amount = 0
        creditors.update_one({"Name": name}, {"$set": {"60 Days": count}})
    if document["30 Days"] > 0 and float(amount) > 0:
        initial = document["30 Days"]
        count = float(initial) - float(amount)
        if count < 0:
            amount = abs(count)
            count = 0
        else:
            amount = 0
        creditors.update_one({"Name": name}, {"$set": {"30 Days": count}})
    return render_template("Creditors/creditors.html")


# Displaying page for creditors report
@app.route("/viewCreditors", methods=['POST'])
def viewCreditors():
    creditors = mongo_client.db.creditors

    list_names = []
    names = creditors.find({}, {"Name": 1, "_id": 0})
    for i in names:
        list_names.append(i['Name'])
    print(list_names)

    list_30 = []
    for i in creditors.find({}, {"30 Days": 1, "_id": 0}):
        list_30.append(i['30 Days'])
    print(list_30)

    list_60 = []
    for i in creditors.find({}, {"60 Days": 1, "_id": 0}):
        list_60.append(i['60 Days'])
    print(list_60)

    list_90 = []
    for i in creditors.find({}, {"90 Days": 1, "_id": 0}):
        list_90.append(i['90 Days'])
    print(list_90)

    list_90_onwards = []
    for i in creditors.find({}, {"90 Onwards": 1, "_id": 0}):
        list_90_onwards.append(i['90 Onwards'])
    print(list_90_onwards)


    data = pd.DataFrame(list(zip(list_names, list_30, list_60, list_90, list_90_onwards)), columns=['Name', '30 Days', '60 Days', '90 Days', '90 Onwards'])

    data['Total'] = data.sum(axis=1)

    file_path = "static/" + str(current_month) + str(current_year) + ".xlsx"
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        workbook.save(file_path)
        print("File Created successfully")

    with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        data.to_excel(writer, 'Creditors')

    table = data.to_html()

    return render_template("Creditors/viewCreditors.html", table=table)


# Debitors Section

# Displaying debitors page
@app.route("/debitors", methods=['GET'])
def debitors():
    return render_template("Debitors/debitors.html")


# Displaying page for adding new debitor
@app.route("/addDebitor", methods=['POST'])
def addDebitor():
    return render_template("Debitors/addDebitor.html")


# Adding new Debitor
@app.route("/aDebitor", methods=['POST'])
def aDebitor():
    name = request.form.get("name")
    date = request.form.get("date")
    amount = request.form.get("amount")
    doc = {
        date: amount
    }
    document = {
        "Name": name,
        "30 Days": float(amount),
        "60 Days": 0,
        "90 Days": 0,
        "90 Onwards": 0,
        "Dates": doc
    }
    debitors = mongo_client.db.debitors
    debitors.insert_one(document)
    return render_template("Debitors/debitors.html")


# Displaying page for adding outstanding
@app.route("/addOutstanding", methods=['POST'])
def addOutstanding():
    debitors = mongo_client.db.debitors
    list_names = []
    names = debitors.find({}, {"Name": 1, "_id": 0})
    for i in names:
        list_names.append(i["Name"])
    return render_template("Debitors/addOutstanding.html", debitors=list_names)


# Adding outstanding amount
@app.route("/aOutstanding", methods=['POST'])
def aOutstanding():
    debitors = mongo_client.db.debitors
    name = request.form.get("name")
    date = request.form.get("date")
    amount = request.form.get("amount")
    document = debitors.find_one({"Name": name})
    doc = document['Dates']

    today = datetime.today().strftime('%Y-%m-%d')
    # print(today)
    date1 = datetime.strptime(date, '%Y-%m-%d')
    date2 = datetime.strptime(today, '%Y-%m-%d')
    difference = (date1 - date2).days
    # difference = abs(date1 - date2)
    difference = abs(difference)
    print(difference)

    if difference <= 30:
        initial = document["30 Days"]
        count = float(initial) + float(amount)
        debitors.update_one({"Name": name}, {"$set": {"30 Days": count}})
    elif difference <= 60:
        initial = document["60 Days"]
        count = float(initial) + float(amount)
        debitors.update_one({"Name": name}, {"$set": {"60 Days": count}})
    elif difference <= 90:
        initial = document["90 Days"]
        count = float(initial) + float(amount)
        debitors.update_one({"Name": name}, {"$set": {"90 Days": count}})
    else:
        initial = document["90 Onwards"]
        count = float(initial) + float(amount)
        debitors.update_one({"Name": name}, {"$set": {"90 Onwards": count}})

    if date in doc:
        initial = float(doc['Production'][date])
        count = initial + float(amount)
        debitors.update_one({"Name": name}, {"$set": {"Dates." + date: count}})
    else:
        debitors.update_one({"Name": name}, {"$set": {"Dates." + date: float(amount)}})

    return render_template("Debitors/debitors.html")


# Displaying page for releasing outstanding
@app.route("/releaseOutstanding", methods=['POST'])
def releaseOutstanding():
    debitors = mongo_client.db.debitors
    list_names = []
    names = debitors.find({}, {"Name": 1, "_id": 0})
    for i in names:
        list_names.append(i["Name"])
    print(list_names)
    return render_template("Debitors/releaseOutstanding.html", debitors=list_names)


# Releasing outstanding amount
@app.route("/rOutstanding", methods=['POST'])
def rOutstanding():
    name = request.form.get("name")
    amount = request.form.get("amount")
    debitors = mongo_client.db.debitors
    document = debitors.find_one({"Name": name})
    print(document)

    if document["90 Onwards"] > 0:
        initial = document["90 Onwards"]
        count = float(initial) - float(amount)
        if count<0:
            amount = abs(count)
            count = 0
        else:
            amount = 0
        debitors.update_one({"Name": name}, {"$set": {"90 Onwards": count}})
    if document["90 Days"] > 0 and float(amount) > 0:
        initial = document["90 Days"]
        count = float(initial) - float(amount)
        if count < 0:
            amount = abs(count)
            count = 0
        else:
            amount = 0
        debitors.update_one({"Name": name}, {"$set": {"90 Days": count}})
    if document["60 Days"] > 0 and float(amount) > 0:
        initial = document["60 Days"]
        count = float(initial) - float(amount)
        if count < 0:
            amount = abs(count)
            count = 0
        else:
            amount = 0
        debitors.update_one({"Name": name}, {"$set": {"60 Days": count}})
    if document["30 Days"] > 0 and float(amount) > 0:
        initial = document["30 Days"]
        count = float(initial) - float(amount)
        if count < 0:
            amount = abs(count)
            count = 0
        else:
            amount = 0
        debitors.update_one({"Name": name}, {"$set": {"30 Days": count}})
    return render_template("Debitors/debitors.html")


# Displaying page for viewing debitors page
@app.route("/viewDebitors", methods=['POST'])
def viewDebitors():
    debitors = mongo_client.db.debitors
    list_names = []
    names = debitors.find({}, {"Name": 1, "_id": 0})
    for i in names:
        list_names.append(i['Name'])
    print(list_names)

    list_30 = []
    for i in debitors.find({}, {"30 Days": 1, "_id": 0}):
        list_30.append(i['30 Days'])
    print(list_30)

    list_60 = []
    for i in debitors.find({}, {"60 Days": 1, "_id": 0}):
        list_60.append(i['60 Days'])
    print(list_60)

    list_90 = []
    for i in debitors.find({}, {"90 Days": 1, "_id": 0}):
        list_90.append(i['90 Days'])
    print(list_90)

    list_90_onwards = []
    for i in debitors.find({}, {"90 Onwards": 1, "_id": 0}):
        list_90_onwards.append(i['90 Onwards'])
    print(list_90_onwards)

    data = pd.DataFrame(list(zip(list_names, list_30, list_60, list_90, list_90_onwards)), columns=['Name', '30 Days', '60 Days', '90 Days', '90 Onwards'])

    data['Total'] = data.sum(axis=1)

    file_path = "static/" + str(current_month) + str(current_year) + ".xlsx"
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        workbook.save(file_path)
        print("File Created successfully")

    with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        data.to_excel(writer, 'Debitors')

    table = data.to_html()

    return render_template("Debitors/viewDebitors.html", table=table)


# Finish Section

@app.route("/finish", methods=['GET'])
def finish():
    return render_template("Finish/finish.html")


@app.route("/addOpening", methods=['POST'])
def addOpening():
    items = mongo_client.db.prodItems
    list_names = []
    for i in items.find({}, {"Name": 1, "_id": 0}):
        list_names.append(i['Name'])

    return render_template("Finish/addOpening.html", items=list_names)


@app.route("/addO", methods=['POST'])
def addO():
    item = request.form.get("item")
    quantity = request.form.get("quantity")
    opening = mongo_client.db.opening
    document = {
        "Name": item,
        "Opening": quantity
    }
    opening.insert_one(document)
    return render_template("Finish/finish.html")


@app.route("/viewFinish", methods=['POST'])
def viewFinish():
    items = mongo_client.db.prodItems
    production = mongo_client.db.production
    dispatch = mongo_client.db.dispatch
    opening = mongo_client.db.opening

    list_names = []
    for i in items.find({}, {"Name": 1, "_id": 0}):
        list_names.append(i['Name'])
    print(list_names)

    list_opening = []
    for i in list_names:
        document = opening.find_one({"Name": i})
        if document:
            list_opening.append(float(document["Opening"]))
        else:
            list_opening.append(0)
    print(list_opening)

    list_production = []
    for i in list_names:
        total = 0
        document = production.find_one({"Name": i})
        if document:
            pro_dic = document["Production"]
            # print(pro_dic)
            for j in pro_dic:
                total += float(pro_dic[j])
            list_production.append(total)
        else:
            list_production.append(0)
    print(list_production)

    list_dispatch = []
    for i in list_names:
        total = 0
        document = dispatch.find_one({"Name": i})
        if document:
            dis_dic = document["Dispatch"]
            # print(dis_dic)
            for j in dis_dic:
                total += float(dis_dic[j])
            list_dispatch.append(total)
        else:
            list_dispatch.append(0)
    print(list_dispatch)

    rates = []
    for i in list_names:
        document = items.find_one({"Name": i})
        if document:
            rates.append(float(document["Rate"]))
        else:
            rates.append(0)
    print(rates)

    data = pd.DataFrame(list(zip(list_names, list_opening, list_production, list_dispatch, rates)), columns=['Item Description', 'Opening', 'Production', 'Dispatched', 'Rate'])

    data['Opening Value'] = data['Opening']*data['Rate']
    data['Production Value'] = data['Production']*data['Rate']
    data['Dispatched Value'] = data['Dispatched']*data['Rate']

    row_sum = data.sum(numeric_only=True)
    # print(row_sum)
    # data.loc['Total'] = data.sum()
    data = data.append(row_sum, ignore_index=True)

    data.at[len(data)-1, 'Item Description'] = 'Total'
    data.at[len(data)-1, 'Rate'] = 0

    # data = data.set_index("Item Description", drop=True)
    print(data.at[len(data)-1, 'Item Description'])

    file_path = "static/" + str(current_month) + str(current_year) + ".xlsx"
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        workbook.save(file_path)
        print("File Created successfully")

    with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        data.to_excel(writer, 'Finish')

    table = data.to_html()

    return render_template("Finish/viewFinish.html", table=table)


@app.route("/analysis", methods=['GET'])
def analysis():
    return render_template("Analysis/analysis.html")


# Analysis Graphs

@app.route("/production_graph")
def production_graph():
    production = mongo_client.db.production
    documents = list(production.find())

    names = [doc['Name'] for doc in documents]
    quantity = [doc['Kgs'] for doc in documents]
    print(names)
    print(quantity)

    fig, ax = plt.subplots()
    plt.bar(names, quantity, color='blue')
    plt.xlabel("Names")
    plt.ylabel("Quantity in Kgs")
    plt.title("Production")
    canvas = FigureCanvas(fig)
    img = BytesIO()
    fig.savefig(img)
    img.seek(0)
    return send_file(img, mimetype='image/png')


@app.route("/dispatch_graph")
def dispatch_graph():
    dispatch = mongo_client.db.dispatch
    items = mongo_client.db.prodItems

    list_names = []
    documents = items.find({}, {"Name": 1, "_id": 0})
    for i in documents:
        list_names.append(i['Name'])
    print(list_names)

    list_quantity = []
    for i in list_names:
        document = dispatch.find_one({"Name": i})
        dis_dic = document['Dispatch']
        total = 0
        for j in dis_dic:
            total += float(dis_dic[j])
        list_quantity.append(total)
    print(list_quantity)

    fig, ax = plt.subplots()
    plt.bar(list_names, list_quantity, color='red')
    plt.xlabel("Names")
    plt.ylabel("Quantity in Kgs")
    plt.title("Dispatch")
    canvas = FigureCanvas(fig)
    img = BytesIO()
    fig.savefig(img)
    img.seek(0)
    return send_file(img, mimetype='image/png')


@app.route("/cnc_report", methods=['GET'])
def cnc_report():
    items = mongo_client.db.prodItems
    production = mongo_client.db.production

    list_names = []
    list_open = []
    for i in items.find({}, {"Name": 1, "_id": 0, "Opening": 1}):
        list_names.append(i['Name'])
        list_open.append(i['Opening'])
    print(list_names)
    print(list_open)

    list_received = []
    for i in list_names:
        dic = production.find_one({"Name": i})
        print(dic)


    return render_template("cnc_report.html")


@app.route("/downloadC", methods=['POST', 'GET'])
def downloadC():
    file_path = str(current_month) + str(current_year) + ".xlsx"
    try:
        return send_from_directory(directory='static', filename=file_path, as_attachment=True)
    except Exception as e:
        return str(e)


@app.route("/downloadP", methods=['POST', 'GET'])
def downloadP():
    today = datetime.datetime.today()
    first = today.replace(day=1)
    prev_month = first - datetime.timedelta(days=1)
    file_path = prev_month.strftime("%B%Y") + ".xlsx"
    try:
        return send_from_directory(directory='static', filename=file_path, as_attachment=True)
    except Exception as e:
        return str(e)



@app.route("/index")
def index():
    client = MongoClient()
    db = client.ME

    if 'attendance' not in db.list_collection_names():
        db.create_collection('attendance')
    if 'creditors' not in db.list_collection_names():
        db.create_collection('creditors')
    if 'debitors' not in db.list_collection_names():
        db.create_collection('debitors')
    if 'dispatch' not in db.list_collection_names():
        db.create_collection('dispatch')
    if 'employees' not in db.list_collection_names():
        db.create_collection('employees')
    if 'opening' not in db.list_collection_names():
        db.create_collection('opening')
    if 'prodItems' not in db.list_collection_names():
        db.create_collection('prodItems')
    if 'production' not in db.list_collection_names():
        db.create_collection('production')
    if 'raw_material' not in db.list_collection_names():
        db.create_collection('raw_material')
    if 'schedule' not in db.list_collection_names():
        db.create_collection('schedule')

    today = datetime.datetime.now().date()

    if (today.month != 12 and today.day == (datetime.date(today.year, today.month + 1, 1) - datetime.timedelta(days=1)).day) or (today.month == 12 and today.day == 31):
        # Today is the last day of the month

        # Download the excel sheet
        db.prodItems.delete_many({})


    if 'username' in session:
        return redirect(url_for('dashboard'))


    # return render_template('index.html')
    return render_template("index.html")


@app.route("/", methods=['POST', 'GET'])
def login():
    return render_template("login.html")


@app.route("/login", methods=['POST', 'GET'])
def login_next():
    if request.method == 'POST':
        users = mongo_client.db.users
        login_user = users.find_one({"Username": request.form["username"]})
        print("User exists")
        if login_user:
            if request.form["password"] == login_user["Password"]:
                session["username"] = request.form["username"]
                return redirect('/index')
            return "Invalid username/password"

        return "Invalid username/password"
    else:
        return redirect("/")


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        users = mongo_client.db.users
        existing_user = users.find_one({"Username": request.form["username"]})

        if existing_user is None:
            users.insert_one({
                "Username": request.form["username"],
                "Password": request.form["password"]
            })
            session["username"] = request.form["username"]
            # mongo.db.createCollection(request.form["username"])
            return redirect("/")
        return "The username already exists"

    return render_template("register.html")


@app.route('/dashboard')
def dashboard():
    if 'username' not in session:
        return redirect(url_for('index'))
    return render_template('index.html', username=session['username'])
#
#
@app.route("/logout")
def logout():
    session.pop("username", None)
    return redirect("/")
#


# def ui(location): #Initiate PyQT5 app
#     qt_app = QApplication(sys.argv)
#     web = QWebEngineView()
#     web.setWindowTitle("Window Name") #Rename to change your window name.
#     # ^ This cannot change between pages
#     web.resize(900, 800) # Set a size
#     web.setZoomFactor(1.5) # Enlarge your content to fit screen
#     web.load(QUrl(location)) #Load Home page at startup
#     web.show() #Show the window
#     sys.exit(qt_app.exec_())


if __name__ == '__main__':
    # Timer(1, lambda: ui("http://127.0.0.1:5000/")).start()
    app.run()
    # ui.run()


# 2023-01-03

# Tasks to do on system reset

# Clear Databases
# 1. Production
# 2. Dispatch

# Creating Login system